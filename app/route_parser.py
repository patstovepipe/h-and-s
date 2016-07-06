# parser.py

from export import Export
import sys
from openpyxl import load_workbook
import getopt
import string
import time
from xlrd import open_workbook

# 'trip' represents a single street and it's house numbers.
suffix_comma = ['st,', 'street,','rd,', 'road,', 'dr,', 'drive,', 'ave,', 'avenue,', 'pl,', 'place,', 'tce,', 'lane,', 'ln,', 'park,', 'prk,',]
suffix = ['st', 'street','rd', 'road', 'dr', 'drive', 'ave', 'avenue', 'pl', 'place', 'tce', 'lane', 'ln', 'park', 'prk',]

delimiters = [',', '/']
oddcode = ['odd', '#odd', 'odd#', 'odd,', '#odd,', 'odd#,']
evencode = ['even#', '#even', 'even', 'even,', '#even,', 'even#,',]

# Importing filename from command line, and collecting the data
def init(route = None, filename = None, username = None, password = None):
    if route != None:
        parsedroute = parse(route)
        print "[route_parser] Parsed the test route: \n", parsedroute
    elif filename != None:
        wb = open_workbook(filename)
        s = wb.sheet_by_index(0)

        # Find data row and data col in spreadsheet
        datarow = None
        datacol = None
        for row_idx in range(0, s.nrows):
            for col_idx in range(0, s.ncols):
                if s.cell(row_idx, col_idx).value and s.cell(row_idx, col_idx).value == "Route Description":
                    datarow = row_idx + 1
                    datacol = col_idx
                    notification = "(Zero indexed) Found datarow at %s and datacol at %s" % (datarow , datacol)
                    if datacol == 4:
                        print notification
                    else:
                        print notification + " but expected datacol at 4"
                    break
            else:
                continue
            break

        if datarow != None and datacol != None:
            parsedroutes = []
            for row_idx in range(datarow, s.nrows):
                cell_obj = s.cell(row_idx, datacol)
                # print cell_obj.value
                parsedroute  = parse(cell_obj.value)
                print parsedroute
                parsedroutes.extend(parsedroute)

            ex = Export()
            ex.set_login_details(username, password)
            ex.set_parsed_routes(parsedroutes)
            ex.generate_excel()

            print "[route_parser] Successfully parsed route description and wrote data to file."
        else:
            print "Route Description column not found, can't proceed."

    return

# parse_trips()
#
# 1. Determine the location of the suffixes rd, ave, etc for each trip
# 2. Determine when the next comma or forward slash (delimiter) occurs
# 3. Using the next delimiter occurence, seperate the words into
#    seperate lists for each trip.
#
#	 Commas often seperate individual house numbers from ranges of house
#	 numbers, so using them exclusively as trip delimiters won't work.
#	 An example would be '...101-141, 151 River St, ...'
#	 Instead, we find the occurence of a suffix, which occurs once per
# 	 trip, and look ahead for the next delimiter. Everything behind the
#	 next delimiter is put in an unordered list, which is handled later.
#
# Suppose we input this route into the above parser:
# 4-40 Beach Dr Even#, 650-776 Mountjoy Ave Even#, 2019-2027 Runnymede Ave Odd# (19)
# The output will look like this:
# [['4-40', 'beach', 'dr', 'even#,'], ['650-776', 'mountjoy', 'ave', 'even#,'], ['2019-2027', 'runnymede', 'ave', 'odd#'], ['(19)']]
def parse_trips(str):
    str = str.lower()
    words = str.split()
    output = []
    endoftrip = 0
    print words
    while len(words) >= 1:
        for word in words:
            if word in suffix_comma:
                endoftrip = words.index(word) + 1
                break
            elif word in suffix:
                for i in range(words.index(word), len(words)):
                    if words[i][-1] in delimiters:
                        endoftrip = i + 1
                        break
                break
        output.append(words[:endoftrip])
        words = words[endoftrip:]
    return output

# organize_trip()
#
#    Takes one list of assorted details, including street name, suffix, house
#	 numbers and a string that specifies odd or even for those numbers.
#    See the comments for parse_trips.
#
# 1. Create a new list and make the first item a string of the street name.
# 2. Generate a list of ints that match the conditions in the list of details,
#    and make that list the second item of the new list.
# 3. Return the new list.
def organize_trip(lst):
    if len(lst) <= 1:
        return []
    streetnameposlist = []
    housenumberlist = []
    housenumbers = []

    print lst

    for word in lst:
        if word in suffix:
            streetnameposlist.append(lst.index(word)-1)
            streetnameposlist.append(lst.index(word))
        if word[0] in string.digits:
            housenumberlist.append(word)
    print streetnameposlist
    streetname = lst[streetnameposlist[0]] + " " + lst[streetnameposlist[1]]
    for word in housenumberlist:
        if '-' in word:
            boundarylist = word.split('-')
            for i in range(int(boundarylist[0]), int(boundarylist[1]), 2):
                housenumbers.append(i)
        else:
            housenumbers.append(int(word))
    output = [streetname, housenumbers]
    return output

def parse(route):
    parsed_items = []
    trips = parse_trips(route)
    for each in trips:
        parsed_items.append(organize_trip(each))
    return parsed_items

if __name__=="__main__":
    init(None, None, None, None)
